import openpyxl
import argparse

parser = argparse.ArgumentParser(
                    prog='ProgramName',
                    description='What the program does',
                    epilog='Text at the bottom of help')


parser.add_argument('-i' , '--infile')           # positional argument
parser.add_argument('-o' , '--outfile')           # positional argument


args = parser.parse_args()

print(args.infile)
print(args.outfile)
datafile = args.infile

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(datafile)
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values

data = []
satir_bos_index = 0
for row in range(0, dataframe1.max_row):
    satir_bos_index += 1
    counter = 0
    row_arr = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if counter == 1:
            temp = col[row].value
            if temp is None:
                continue
            else:
                temp = temp.replace(' ', '')
                if temp == '':
                    continue
                else:
                    row_arr.append(col[row].value)
        else:
            row_arr.append(col[row].value)
        counter += 1
        if len(row_arr) == 9:
            data.append(row_arr)


maximum = 0
index_counter = {0: ''}
for i in range(1, len(data)):
    first_col = str(data[i][0])
    first_col = first_col.replace(',', '.')
    first_coll_arr = first_col.split('.')
    if len(first_coll_arr) > maximum:
        maximum = len(first_coll_arr)
    index_counter[i] = len(first_coll_arr)

for i in range(0, len(data)):
    for j in range(maximum):
        data[i].insert(1 +j, index_counter[i])

from openpyxl import load_workbook, Workbook

wb = Workbook()
ws = wb.active

for row in data:
    ws.append(row)

wb.save(args.outfile)